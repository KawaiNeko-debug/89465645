import os
import sys
import tempfile
import types
import unittest
from unittest.mock import patch


for key, value in {
    "BASE_URL": "https://activity.example.test",
    "PASSPORT_URL": "https://passport.example.test",
    "REFERER": "https://activity.example.test/",
    "VOTE_CAMPAIGN_URL": "https://campaign.example.test/portal/brand-campaign",
    "VOTE_API_BASE": "https://campaign.example.test",
    "SLIDER_ID": "slider",
    "WRAPPER_ID": "wrapper",
    "HEADER_CLIENT_TYPE": "client-type",
    "HEADER_ACCESS_TOKEN": "access-token",
    "TOKEN_KEY": "token-key",
}.items():
    os.environ.setdefault(key, value)

if "fake_useragent" not in sys.modules:
    fake_useragent = types.ModuleType("fake_useragent")
    fake_useragent.UserAgent = type("UserAgent", (), {"random": "test-agent"})
    sys.modules["fake_useragent"] = fake_useragent

if "playwright.sync_api" not in sys.modules:
    playwright = types.ModuleType("playwright")
    sync_api = types.ModuleType("playwright.sync_api")
    sync_api.sync_playwright = lambda: None
    sync_api.Page = object
    sync_api.TimeoutError = TimeoutError
    playwright.sync_api = sync_api
    sys.modules["playwright"] = playwright
    sys.modules["playwright.sync_api"] = sync_api

from openpyxl import load_workbook

from h3.campaign_vote import (
    CAMPAIGN_URL,
    VOTE_ACTIVITY_ACCESS_ID,
    VOTE_PRODUCT_NAME,
    VOTE_PRODUCT_SKU,
    campaign_session_ready,
    inspect_vote_config,
    is_vote_date,
    vote_environment_error,
)
from h3.feature_flags import SECKILL_ENABLED
from h3.merge_results import pick_result
from h3.report import build_summary, merge_records_with_expected, write_xlsx
from h3.script import ApiClient, should_retry


def vote_config(*, can_vote=True, my_sku="", product_name=VOTE_PRODUCT_NAME):
    return {
        "success": True,
        "code": 200,
        "data": {
            "activityStatus": 2,
            "myVotedProductSku": my_sku or None,
            "voteProductConfigList": [
                {
                    "productSku": VOTE_PRODUCT_SKU,
                    "productName": product_name,
                    "canVote": can_vote,
                }
            ],
        },
    }


class CampaignVoteTests(unittest.TestCase):
    def test_seckill_feature_is_disabled_and_runtime_skips_requests(self):
        self.assertFalse(SECKILL_ENABLED)

        client = ApiClient("token", "secret", 1, object(), user_agent="test-agent")
        events = []
        client.fetch_brand_activity_config = lambda: events.append("config") or {}
        client.get_lottery_activity_code = lambda config: events.append("lottery_code") or "LOTTERY"
        client.get_seckill_category_ids = lambda config: self.fail("关闭时不应解析秒杀分类")
        client.fetch_seckill_records = lambda *args, **kwargs: self.fail("关闭时不应请求秒杀记录")

        def fetch_component(tag, fetcher, success_attr):
            events.append(tag)
            setattr(client, success_attr, True)
            return []

        client.fetch_activity_component_with_retry = fetch_component
        result = client.fetch_activity_records()

        self.assertEqual(events, ["config", "lottery_code", "抽奖记录"])
        self.assertEqual(result, {"seckill": [], "lottery": []})
        self.assertTrue(client.activity_fetch_success)

    def test_vote_endpoints_must_come_from_valid_environment_values(self):
        error = vote_environment_error("", "")
        self.assertIn("VOTE_CAMPAIGN_URL", error)
        self.assertIn("VOTE_API_BASE", error)
        self.assertIn("VOTE_API_BASE", vote_environment_error(CAMPAIGN_URL, "invalid"))
        self.assertEqual(vote_environment_error(CAMPAIGN_URL, os.environ["VOTE_API_BASE"]), "")

    def test_vote_window_includes_all_three_requested_days(self):
        self.assertFalse(is_vote_date("2026-07-28"))
        self.assertTrue(is_vote_date("2026-07-29 08:00:00"))
        self.assertTrue(is_vote_date("2026-07-30"))
        self.assertTrue(is_vote_date("2026-07-31 23:59:59"))
        self.assertFalse(is_vote_date("2026-08-01"))

    def test_ready_and_already_states_are_distinct(self):
        self.assertEqual(inspect_vote_config(vote_config())["state"], "ready")
        already = inspect_vote_config(vote_config(can_vote=False, my_sku=VOTE_PRODUCT_SKU))
        self.assertEqual(already["state"], "already")
        self.assertEqual(already["my_sku"], VOTE_PRODUCT_SKU)

    def test_campaign_session_requires_authenticated_customer(self):
        self.assertTrue(campaign_session_ready({"code": 200, "body": {"customerCode": "test"}}))
        self.assertFalse(campaign_session_ready({"code": 200, "body": {}}))
        self.assertFalse(campaign_session_ready({"code": 401, "body": {"customerCode": "test"}}))

    def test_vote_orchestration_opens_sso_page_then_submits_and_verifies(self):
        events = []

        class FakePage:
            def goto(self, url, **kwargs):
                events.append(("goto", url, kwargs))

        client = ApiClient("token", "secret", 1, FakePage(), user_agent="test-agent")
        client.vote_required = True
        configs = iter(
            [
                vote_config(can_vote=True),
                vote_config(can_vote=False, my_sku=VOTE_PRODUCT_SKU),
            ]
        )

        def session_ready(quiet=False):
            events.append(("session",))
            return True

        def fetch_config():
            events.append(("config",))
            return next(configs)

        def browser_request(method, url, **kwargs):
            events.append(("submit", method, url, kwargs.get("payload")))
            return {"success": True, "code": 200, "data": None}

        client._campaign_session_ready = session_ready
        client._fetch_vote_config = fetch_config
        client._browser_fetch_json_once = browser_request

        with patch("h3.script.time.sleep", return_value=None):
            self.assertTrue(client.execute_campaign_vote())

        self.assertEqual([event[0] for event in events], ["goto", "session", "config", "submit", "config"])
        self.assertEqual(events[0][1], CAMPAIGN_URL)
        self.assertEqual(
            events[3][3],
            {"activityAccessId": VOTE_ACTIVITY_ACCESS_ID, "productSku": VOTE_PRODUCT_SKU},
        )
        self.assertTrue(client.vote_success)
        self.assertIn(VOTE_PRODUCT_NAME, client.vote_status)

    def test_initial_401_triggers_sso_then_continues_without_page_reload(self):
        events = []

        class FakePage:
            def goto(self, url, **kwargs):
                events.append(("goto", url))

        client = ApiClient("token", "secret", 1, FakePage(), user_agent="test-agent")
        client.vote_required = True
        readiness = iter([False, True])
        configs = iter(
            [
                vote_config(can_vote=True),
                vote_config(can_vote=False, my_sku=VOTE_PRODUCT_SKU),
            ]
        )

        def session_ready(quiet=False):
            events.append(("session", quiet))
            return next(readiness)

        client._campaign_session_ready = session_ready
        client._trigger_campaign_sso = lambda: events.append(("trigger_sso",)) or True
        client._fetch_vote_config = lambda: events.append(("config",)) or next(configs)
        client._browser_fetch_json_once = (
            lambda method, url, **kwargs: events.append(("submit", method, url))
            or {"success": True, "code": 200}
        )

        with patch("h3.script.time.sleep", return_value=None):
            self.assertTrue(client.execute_campaign_vote())

        self.assertEqual(
            [event[0] for event in events],
            ["goto", "session", "trigger_sso", "session", "config", "submit", "config"],
        )
        self.assertEqual(sum(1 for event in events if event[0] == "goto"), 1)

    def test_sso_failure_never_calls_vote_apis(self):
        events = []

        class FakePage:
            def goto(self, url, **kwargs):
                events.append(("goto", url))

        client = ApiClient("token", "secret", 1, FakePage(), user_agent="test-agent")
        client.vote_required = True
        client._campaign_session_ready = lambda quiet=False: events.append(("session", quiet)) or False
        client._trigger_campaign_sso = lambda: events.append(("trigger_sso",)) or True
        client._fetch_vote_config = lambda: self.fail("SSO 失败时不应查询投票配置")
        client._browser_fetch_json_once = lambda *args, **kwargs: self.fail("SSO 失败时不应提交投票")

        with patch("h3.script.time.sleep", return_value=None):
            self.assertFalse(client.execute_campaign_vote())

        self.assertEqual(sum(1 for event in events if event[0] == "goto"), 1)
        self.assertEqual(sum(1 for event in events if event[0] == "trigger_sso"), 1)
        self.assertIn("SSO 会话未建立", client.vote_status)

    def test_vote_failure_does_not_repeat_the_full_sign_flow(self):
        result = {
            "sign_success": True,
            "password_error": False,
            "risk_controlled": False,
            "banned_account": False,
            "vote_required": True,
            "vote_success": False,
        }
        self.assertFalse(should_retry(result))

    def test_refuses_other_locked_product_or_name_mismatch(self):
        conflict = inspect_vote_config(vote_config(my_sku="OTHER-SKU"))
        self.assertEqual(conflict["state"], "conflict")
        mismatch = inspect_vote_config(vote_config(product_name="错误商品"))
        self.assertEqual(mismatch["state"], "error")

    def test_submit_success_is_not_enough_without_target_vote_verification(self):
        class FakePage:
            def goto(self, url, **kwargs):
                return None

        client = ApiClient("token", "secret", 1, FakePage(), user_agent="test-agent")
        client.vote_required = True
        client._campaign_session_ready = lambda quiet=False: True
        client._fetch_vote_config = lambda: vote_config(can_vote=True)
        client._browser_fetch_json_once = lambda *args, **kwargs: {"success": True, "code": 200}

        with patch("h3.script.time.sleep", return_value=None):
            self.assertFalse(client.execute_campaign_vote())

        self.assertFalse(client.vote_success)
        self.assertIn("复核未确认", client.vote_status)

    def test_retry_merge_keeps_successful_vote(self):
        initial = {
            "sign_success": True,
            "data_fetch_completed": True,
            "points_fetch_success": True,
            "activity_fetch_success": True,
            "vote_required": True,
            "vote_success": False,
            "vote_status": "投票失败",
            "retry_count": 0,
        }
        retry = dict(initial, vote_success=True, vote_status="今日已投票", retry_count=1)
        picked = pick_result(initial, retry)
        self.assertTrue(picked["vote_success"])
        self.assertEqual(picked["vote_status"], "今日已投票")

    def test_xlsx_contains_daily_vote_columns(self):
        record = {
            "account_index": 1,
            "username": "test-account",
            "group_position": "1组账号1",
            "sign_success": True,
            "sign_status": "签到成功",
            "final_points": 100,
            "vote_required": True,
            "vote_success": True,
            "vote_status": f"投票成功：{VOTE_PRODUCT_NAME}",
            "vote_time": "2026-07-29 08:30:00",
            "activity_records": {
                "seckill": [{"title": "旧秒杀奖品", "claimed": True}],
                "lottery": [{"title": "抽奖奖品", "claimed": True}],
            },
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [record])
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet["I1"].value, "当天投票情况")
            self.assertEqual(sheet["J1"].value, "投票时间")
            self.assertIn(VOTE_PRODUCT_NAME, sheet["I2"].value)
            self.assertEqual(sheet["J2"].value, "2026-07-29 08:30:00")
            headers = [cell.value for cell in sheet[1]]
            values = [cell.value for row in sheet.iter_rows() for cell in row]
            self.assertNotIn("秒杀一", headers)
            self.assertNotIn("秒杀二", headers)
            self.assertNotIn("旧秒杀奖品", values)
            self.assertEqual(sheet["K1"].value, "抽奖一")
            self.assertEqual(sheet["K2"].value, "抽奖奖品")
            workbook.close()

    def test_missing_account_is_counted_as_required_vote_on_campaign_date(self):
        records = merge_records_with_expected([], {(1, 1): "missing-account"}, "2026-07-29")
        self.assertEqual(len(records), 1)
        self.assertTrue(records[0]["vote_required"])
        self.assertEqual(records[0]["vote_status"], "缺少投票结果")

        summary = build_summary(records, expected_total=1)
        self.assertEqual(summary["vote_required"], 1)
        self.assertEqual(summary["vote_success"], 0)
        self.assertEqual(summary["problem_count"], 1)


if __name__ == "__main__":
    unittest.main()
