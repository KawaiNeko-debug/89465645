import json
import os
import sys
import tempfile
import unittest
import io
import zipfile
from datetime import datetime
from types import SimpleNamespace
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from h3.account_data import (
    browser_fetch_json,
    empty_account_data,
    normalize_coupon,
    parse_coupon_response,
    parse_invoice_profile_exists,
    parse_invoice_statistics,
    parse_invoice_order_page,
    predict_pcb_smt,
    sum_pcb_invoice_orders,
)
from h3.merge_results import pick_result
from h3.listing_gift import inspect_listing_gift_response, is_listing_gift_date
from h3.report import max_lottery_count, normalize_activity_records, write_xlsx
from h3.retry_components import build_retry_matrix, component_status, retry_components
from h3.runner_recovery import should_rerun
from h3.schedule_guard import has_manual_run_today
from h3.exchange_history import exchange_status_text, normalize_exchange_records
from h3.campaign_vote import (
    activity_config_payload,
    can_vote_after_sign,
    inspect_vote_config,
    is_vote_date,
    parse_lottery_winning_response,
)
from h3.dynamic_groups import (
    compact_json,
    configured_groups,
    dispatch_once,
    download_groups,
    new_chain_state,
    encode_chain_state,
    load_chain_state,
)
from h3.category_reports import main as category_reports_main
from h3.report import (
    build_message,
    is_problem_record,
    load_account_lookup,
    load_credential_lookup,
    mask_account,
    normalize_record,
    redact_accounts_for_log,
    report_password,
    resolve_output_xlsx_path,
)
from h3.test_report import prepare_manifest


def lottery_rows(count: int) -> list[dict]:
    return [
        {
            "title": f"奖品{index}",
            "claimed": index % 2 == 0,
            "status_text": "已经领取" if index % 2 == 0 else "未领取",
        }
        for index in range(1, count + 1)
    ]


def account_data_with_amount(within=100, over=0) -> dict:
    data = empty_account_data()
    data.update(
        {
            "fetch_success": True,
            "invoice_fetch_success": True,
            "coupon_fetch_success": True,
            "invoice_profile_status": "有",
            "invoice_profile_exists": True,
            "invoice_month_threshold": 12,
            "invoice_within_months_amount": within,
            "invoice_over_months_amount": over,
            "pcb_order_fetch_success": True,
            "pcb_within_months_amount": within,
            "pcb_over_months_amount": over,
            "pcb_total_amount": within + over,
            "pcb_amount_shortfall": max(0, 40 - within - over),
        }
    )
    return data


def record(index: int, lottery_count: int, account_data=None) -> dict:
    return {
        "account_index": index,
        "username": f"account{index}",
        "group_number": 1,
        "group_position": f"1组账号{index}",
        "sign_success": True,
        "sign_status": "签到成功",
        "final_points": 100,
        "activity_records": {"seckill": [], "lottery": lottery_rows(lottery_count)},
        "account_data_required": True,
        "account_data_fetch_success": True,
        "account_data": account_data or account_data_with_amount(),
        "listing_gift_required": True,
        "listing_gift_success": True,
        "listing_gift_attempted": True,
        "listing_gift_status": "上市礼包领取成功",
        "listing_gift_time": "2026-08-05 12:00:00",
        "listing_gift_detail": "上市礼包领取成功",
    }


class DynamicLotteryTests(unittest.TestCase):
    def test_activity_config_request_never_uses_an_empty_payload(self):
        self.assertEqual(
            activity_config_payload(" campaign-id "),
            {"activityAccessId": "campaign-id"},
        )
        self.assertEqual(activity_config_payload(""), {})

    def test_normalization_never_truncates_supported_counts(self):
        for count in (0, 1, 3, 5, 8, 15):
            with self.subTest(count=count):
                normalized = normalize_activity_records(
                    {"seckill": [], "lottery": lottery_rows(count)}
                )
                self.assertEqual(len(normalized["lottery"]), count)

    def test_xlsx_headers_follow_the_largest_lottery_count(self):
        for count in (0, 1, 3, 5, 8, 15):
            with self.subTest(count=count), tempfile.TemporaryDirectory() as temp_dir:
                path = os.path.join(temp_dir, "report.xlsx")
                write_xlsx(path, [record(1, count)])
                sheet = load_workbook(path)["签到汇总"]
                headers = [cell.value for cell in sheet[1]]
                self.assertEqual(
                    [value for value in headers if str(value or "").startswith("抽奖")],
                    [f"抽奖{index}" for index in range(1, count + 1)],
                )
                self.assertEqual(
                    [value for value in headers if str(value or "").startswith("领取情况")],
                    [f"领取情况{index}" for index in range(1, count + 1)],
                )

    def test_accounts_are_padded_to_the_daily_maximum(self):
        rows = [record(1, 2), record(2, 6), record(3, 8)]
        self.assertEqual(max_lottery_count(rows), 8)
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, rows)
            sheet = load_workbook(path)["签到汇总"]
            headers = [cell.value for cell in sheet[1]]
            first_row = next(row for row in sheet.iter_rows(min_row=2) if row[2].value == "account1")
            self.assertEqual(headers.count("抽奖8"), 1)
            self.assertIsNone(first_row[headers.index("抽奖3")].value)
            self.assertIsNone(first_row[headers.index("抽奖8")].value)

    def test_retry_merge_keeps_the_more_complete_lottery_array(self):
        initial = record(1, 15)
        initial.update({"sign_success": False, "activity_fetch_success": True, "retry_count": 0})
        retry = record(1, 3)
        retry.update({"sign_success": True, "activity_fetch_success": True, "retry_count": 1})
        picked = pick_result(initial, retry)
        self.assertEqual(len(picked["activity_records"]["lottery"]), 15)


class ExchangeHistoryTests(unittest.TestCase):
    def test_status_mapping_and_august_2026_cutoff(self):
        rows = [
            {
                "integralChangeType": 5,
                "exchangeStates": state,
                "exchangeNum": 1,
                "integralChangeNum": 100,
                "createTime": f"2026-08-0{state}T02:04:01.000Z",
                "goodsName": f"奖品{state}",
            }
            for state in range(1, 7)
        ]
        rows.append(
            {
                "integralChangeType": 5,
                "exchangeStates": 6,
                "createTime": "2026-07-31T15:59:59.000Z",
                "goodsName": "七月奖品",
            }
        )
        normalized = normalize_exchange_records(rows)
        self.assertEqual([item["status_text"] for item in normalized], [
            "已确认收货", "已退回", "已冻结", "已兑换待发货", "已发货", "已兑换"
        ])
        self.assertNotIn("七月奖品", [item["title"] for item in normalized])
        self.assertEqual(normalized[-1]["created_at"], "2026-08-01 10:04:01")
        self.assertEqual(exchange_status_text({"integralChangeType": 2}), "已发到账户")

    def test_xlsx_has_unique_prize_pairs_and_status_colors(self):
        first = record(1, 0)
        first["activity_records"]["exchange"] = [
            {"title": "抽纸", "status": 1, "status_text": "已兑换", "quantity": 1, "points": 390, "created_at": "2026-08-06 10:04:01"},
            {"title": "水杯", "status": 5, "status_text": "已退回", "quantity": 1, "points": 0, "created_at": "2026-08-05 09:00:00"},
        ]
        second = record(2, 0)
        second["activity_records"]["exchange"] = [
            {"title": "抽纸", "status": 3, "status_text": "已兑换待发货", "quantity": 2, "points": 780, "created_at": "2026-08-08 08:00:00"},
            {"title": "露营车", "status": 6, "status_text": "已确认收货", "quantity": 1, "points": 1000, "created_at": "2026-08-07 08:00:00"},
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [first, second])
            sheet = load_workbook(path)["签到汇总"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(
                [value for value in headers if str(value or "").startswith("兑换物品：")],
                ["兑换物品：抽纸", "兑换物品：水杯", "兑换物品：露营车"],
            )
            paper_status = headers.index("兑换状态：抽纸") + 1
            cup_status = headers.index("兑换状态：水杯") + 1
            cart_status = headers.index("兑换状态：露营车") + 1
            self.assertEqual(sheet.cell(2, paper_status).fill.fgColor.rgb, "00FFD966")
            self.assertEqual(sheet.cell(2, cup_status).fill.fgColor.rgb, "00F8696B")
            self.assertEqual(sheet.cell(3, paper_status).fill.fgColor.rgb, "00C6E0B4")
            self.assertEqual(sheet.cell(3, cart_status).fill.fgColor.rgb, "00C6E0B4")
            paper_detail = headers.index("兑换物品：抽纸") + 1
            self.assertIn("2件", sheet.cell(3, paper_detail).value)

    def test_retry_merge_keeps_longer_exchange_array(self):
        initial = record(1, 1)
        initial.update({"sign_success": False, "activity_fetch_success": True})
        initial["activity_records"]["exchange"] = [
            {"title": f"奖品{index}", "status_text": "已兑换"} for index in range(3)
        ]
        retry = record(1, 2)
        retry.update({"sign_success": True, "activity_fetch_success": True, "retry_count": 1})
        retry["activity_records"]["exchange"] = [{"title": "奖品0", "status_text": "已兑换"}]
        picked = pick_result(initial, retry)
        self.assertEqual(len(picked["activity_records"]["lottery"]), 2)
        self.assertEqual(len(picked["activity_records"]["exchange"]), 3)


class AccountDataTests(unittest.TestCase):
    def test_har_invoice_fields_profile_signal_and_bodyless_post(self):
        response = {
            "success": True,
            "code": 200,
            "data": {
                "vatOrdinaryInvoiceMoney": 300.0,
                "overTimeNum": 12,
                "overTimeElecInvoiceMoney": 0.51,
                "customerInvoiceInfoVO": {"vatCompanyName": "fixture"},
            },
        }
        self.assertEqual(
            parse_invoice_statistics(response),
            {"month_threshold": 12, "within_amount": 300.0, "over_amount": 0.51},
        )
        self.assertTrue(parse_invoice_profile_exists([response]))

        class FakePage:
            def __init__(self):
                self.argument = None

            def evaluate(self, _source, argument):
                self.argument = argument
                return {"ok": True, "data": {"success": True, "code": 200}}

        page = FakePage()
        self.assertTrue(browser_fetch_json(page, "POST", "/path", None)["success"])
        self.assertIsNone(page.argument["payload"])
        browser_fetch_json(page, "POST", "/path", {"invoiceType": "2"}, form_encoded=True)
        self.assertTrue(page.argument["formEncoded"])

    def test_invoice_fields_and_profile_presence(self):
        parsed = parse_invoice_statistics(
            {"success": True, "data": {"overTimeNum": 18, "vatSpecialInvoiceMoney": "12.50", "overTimeInvoiceMoney": 3}}
        )
        self.assertEqual(parsed, {"month_threshold": 18, "within_amount": 12.5, "over_amount": 3.0})
        self.assertTrue(
            parse_invoice_profile_exists(
                [{"success": True, "data": {"customerInvoiceInfoVO": {"invoiceTitle": "测试企业"}}}]
            )
        )
        self.assertFalse(parse_invoice_profile_exists([{"success": True, "data": {}}]))
        self.assertFalse(
            parse_invoice_profile_exists(
                [{"success": True, "data": {"customerInvoiceInfoVO": {"invoiceTitle": ""}}}]
            )
        )

    def test_coupon_status_expiry_and_rules_are_preserved(self):
        response = {
            "success": True,
            "data": {
                "records": [
                    {
                        "couponId": "1",
                        "couponName": "PCB高多层150元优惠券",
                        "businessType": "PCB",
                        "expirationTime": "2026-12-31 23:59:59",
                        "useRule": "满额可用",
                    }
                ]
            },
        }
        coupon = parse_coupon_response(response, "unused")[0]
        self.assertEqual(coupon["status"], "未使用")
        self.assertEqual(coupon["expires_at"], "2026-12-31 23:59:59")
        self.assertEqual(coupon["rule_text"], "满额可用")
        legacy = parse_coupon_response(
            {"success": True, "data": {"root": [{"historyCoupon": {"couponName": "旧版券"}}]}},
            "expired",
        )
        self.assertEqual(legacy[0]["name"], "旧版券")

    def test_har_coupon_date_fields_are_preserved(self):
        coupon = normalize_coupon(
            {
                "couponName": "fixture",
                "startDate": "2026-08-01 00:00:00",
                "endDate": "2026-08-31 23:59:59",
            },
            "unused",
        )
        self.assertEqual(coupon["valid_from"], "2026-08-01 00:00:00")
        self.assertEqual(coupon["expires_at"], "2026-08-31 23:59:59")

    def test_all_prediction_branches(self):
        insufficient = empty_account_data()
        self.assertEqual(predict_pcb_smt(insufficient)[0], "数据不足")

        no_spend = account_data_with_amount(0, 0)
        self.assertEqual(predict_pcb_smt(no_spend)[0], "不可能")

        current = account_data_with_amount()
        current["coupons"]["unused"] = [
            normalize_coupon({"couponName": "PCB+SMT专用券"}, "unused")
        ]
        self.assertEqual(predict_pcb_smt(current)[0], "很小可能")

        expired = account_data_with_amount()
        expired["coupons"]["expired"] = [
            normalize_coupon({"couponName": "PCB与SMT组合券"}, "expired")
        ]
        self.assertEqual(predict_pcb_smt(expired)[0], "很大可能")

        never_seen = account_data_with_amount()
        self.assertEqual(predict_pcb_smt(never_seen)[0], "100%可能")

    def test_coupon_sheets_handle_duplicates_illegal_chars_and_long_names(self):
        first = account_data_with_amount()
        second = account_data_with_amount()
        second.update({"invoice_profile_status": "无", "invoice_profile_exists": False})
        shared_name = "PCB/SMT:*?[]优惠券" + "A" * 30
        colliding_name = "PCB_SMT______优惠券" + "A" * 30
        first["coupons"]["unused"] = [
            normalize_coupon({"couponName": shared_name, "expirationTime": "2026-10-01"}, "unused")
        ]
        second["coupons"]["unused"] = [
            normalize_coupon({"couponName": shared_name, "expirationTime": "2026-11-01"}, "unused"),
            normalize_coupon({"couponName": colliding_name, "expirationTime": "2026-12-01"}, "unused"),
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [record(1, 0, first), record(2, 0, second)])
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames[1], "PCB+SMT券")
            coupon_sheets = workbook.sheetnames[2:]
            self.assertEqual(len(coupon_sheets), 2)
            self.assertTrue(all(len(name) <= 31 for name in coupon_sheets))
            self.assertEqual(len({name.lower() for name in coupon_sheets}), 2)
            self.assertEqual(workbook[coupon_sheets[0]].max_row + workbook[coupon_sheets[1]].max_row, 5)
            for sheet_name in coupon_sheets:
                sheet = workbook[sheet_name]
                headers = [cell.value for cell in sheet[1]]
                self.assertEqual(headers[:5], ["序号", "客编", "密码", "开票资料", "优惠券过期时间"])
                invoice_column = headers.index("开票资料") + 1
                for row_index in range(2, sheet.max_row + 1):
                    expected = "00C00000" if sheet.cell(row_index, invoice_column).value == "无" else "00008000"
                    self.assertEqual(sheet.cell(row_index, invoice_column).font.color.rgb, expected)

    def test_pcb_smt_sheet_is_second_and_contains_credentials(self):
        data = account_data_with_amount()
        data["coupons"]["unused"] = [
            normalize_coupon(
                {"couponName": "PCB+SMT组合券", "expirationTime": "2026-12-31"},
                "unused",
            )
        ]
        row = record(1, 0, data)
        row.update({"username": "customer123", "password": "plain-password"})
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [row])
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames[:2], ["签到汇总", "PCB+SMT券"])
            main_headers = [cell.value for cell in workbook["签到汇总"][1]]
            self.assertEqual(main_headers[2:4], ["客编", "密码"])
            self.assertEqual(workbook["PCB+SMT券"][2][1].value, "customer123")
            self.assertEqual(workbook["PCB+SMT券"][2][2].value, "plain-password")
            pcb_headers = [cell.value for cell in workbook["PCB+SMT券"][1]]
            self.assertEqual(pcb_headers[:6], ["序号", "客编", "密码", "开票资料", "优惠券名称", "优惠券过期时间"])
            self.assertEqual(workbook["PCB+SMT券"][2][3].value, "有")
            self.assertEqual(workbook["PCB+SMT券"][2][3].font.color.rgb, "00008000")

    def test_future_or_expired_pcb_smt_coupon_creates_empty_status_sheet(self):
        data = account_data_with_amount()
        data["coupons"]["unused"] = [
            normalize_coupon(
                {
                    "couponName": "PCB+SMT组合券",
                    "effectiveTime": "2027-01-01",
                    "expirationTime": "2027-12-31",
                },
                "unused",
            )
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [record(1, 0, data)])
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames[:2], ["签到汇总", "PCB+SMT券"])
            self.assertEqual(workbook["PCB+SMT券"]["A2"].value, "当前未检测到可用的PCB+SMT券")

    def test_pcb_smt_sheet_is_always_created_when_coupon_list_is_empty(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [record(1, 0, account_data_with_amount())])
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames[:2], ["签到汇总", "PCB+SMT券"])
            self.assertEqual(workbook["PCB+SMT券"]["A2"].value, "当前未检测到可用的PCB+SMT券")

    def test_no_profile_keeps_invoice_amount_cells_empty(self):
        data = account_data_with_amount(10, 5)
        data.update(
            {
                "invoice_profile_status": "无",
                "invoice_profile_exists": False,
                "invoice_within_months_amount": None,
                "invoice_over_months_amount": None,
            }
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [record(1, 0, data)])
            sheet = load_workbook(path)["签到汇总"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            self.assertEqual(sheet.cell(2, headers["开票资料"]).value, "无")
            self.assertIsNone(sheet.cell(2, headers["不超过12个月可开金额"]).value)
            self.assertIsNone(sheet.cell(2, headers["超过12个月可开金额"]).value)

    def test_pcb_invoice_order_types_amount_fallback_and_deduplication(self):
        response = {
            "data": {
                "pageNum": 1,
                "totalPages": 2,
                "data": [
                    {"businessOrderType": 1, "invoiceMoney": "20.50", "businessOrderCode": "a"},
                    {"businessOrderType": 2, "invoiceMoney": None, "orderMoney": "19.50", "businessOrderCode": "b"},
                    {"businessOrderType": 3, "invoiceMoney": "100", "businessOrderCode": "c"},
                    {"businessOrderType": 9, "invoiceMoney": "100", "businessOrderCode": "d"},
                ],
            }
        }
        page = parse_invoice_order_page(response)
        self.assertEqual(page["total_pages"], 2)
        self.assertEqual([item["business_order_type"] for item in page["orders"]], [1, 2])
        total, count = sum_pcb_invoice_orders([response, response])
        self.assertEqual(total, 40.0)
        self.assertEqual(count, 2)

    def test_pcb_threshold_boundaries(self):
        for amount, expected, shortfall in (
            (0, "不可能", 40),
            (39.99, "不可能", 0.01),
            (40, "100%可能", 0),
            (40.01, "100%可能", 0),
        ):
            with self.subTest(amount=amount):
                data = account_data_with_amount(amount, 0)
                self.assertEqual(predict_pcb_smt(data)[0], expected)
                self.assertAlmostEqual(data["pcb_amount_shortfall"], shortfall, places=2)


class ListingGiftTests(unittest.TestCase):
    def test_date_window_is_exact(self):
        self.assertFalse(is_listing_gift_date("2026-08-04"))
        self.assertTrue(is_listing_gift_date("2026-08-05 23:59:59"))
        self.assertTrue(is_listing_gift_date("2026-08-06"))
        self.assertFalse(is_listing_gift_date("2026-08-07"))

    def test_har_success_shape_is_required(self):
        success = inspect_listing_gift_response(
            {"success": True, "code": 200, "data": {"success": True, "orderCode": "test"}}
        )
        self.assertEqual(success["state"], "received")
        self.assertTrue(success["success"])
        self.assertEqual(success["order_code"], "test")
        self.assertFalse(inspect_listing_gift_response({"success": True, "data": {}})["success"])

    def test_already_received_is_idempotent_success(self):
        result = inspect_listing_gift_response(
            {"success": False, "message": "今日已经领取，请勿重复领取"}
        )
        self.assertEqual(result["state"], "already")
        self.assertTrue(result["success"])

    def test_bodyless_repeat_response_is_idempotent_success(self):
        result = inspect_listing_gift_response(
            {"success": True, "code": 200, "data": {"success": False, "orderCode": None}}
        )
        self.assertEqual(result["state"], "already")
        self.assertTrue(result["success"])

    def test_retry_merge_preserves_successful_gift_result(self):
        initial = record(1, 1)
        initial.update({"sign_success": False, "retry_count": 0})
        retry = record(1, 1)
        retry.update(
            {
                "sign_success": True,
                "retry_count": 1,
                "listing_gift_success": False,
                "listing_gift_status": "领取失败",
            }
        )
        picked = pick_result(initial, retry)
        self.assertTrue(picked["listing_gift_success"])
        self.assertEqual(picked["listing_gift_status"], "上市礼包领取成功")

    def test_xlsx_contains_gift_status_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "report.xlsx")
            write_xlsx(path, [record(1, 1)])
            sheet = load_workbook(path)["签到汇总"]
            headers = [cell.value for cell in sheet[1]]
            column = headers.index("上市礼包领取情况") + 1
            self.assertIn("上市礼包领取成功", sheet.cell(2, column).value)
            self.assertEqual(sheet.cell(2, column).font.color.rgb, "00008000")


class VoteTests(unittest.TestCase):
    def test_log_redacts_active_credentials(self):
        required = {
            "BASE_URL": "https://example.invalid",
            "PASSPORT_URL": "https://example.invalid/login",
            "REFERER": "https://example.invalid",
            "SLIDER_ID": "slider",
            "WRAPPER_ID": "wrapper",
            "HEADER_CLIENT_TYPE": "client",
            "HEADER_ACCESS_TOKEN": "access",
            "TOKEN_KEY": "token",
        }
        with patch.dict(os.environ, required, clear=False):
            import importlib
            script_module = importlib.import_module("h3.script")
        original = set(script_module._SENSITIVE_LOG_VALUES)
        try:
            script_module._SENSITIVE_LOG_VALUES.clear()
            script_module._SENSITIVE_LOG_VALUES.update({"user-secret", "pass-secret"})
            with patch("builtins.print") as printer:
                script_module.log("user-secret pass-secret")
            rendered = str(printer.call_args.args[0])
            self.assertNotIn("user-secret", rendered)
            self.assertNotIn("pass-secret", rendered)
            self.assertEqual(rendered.count("[REDACTED]"), 2)
        finally:
            script_module._SENSITIVE_LOG_VALUES.clear()
            script_module._SENSITIVE_LOG_VALUES.update(original)

    def test_vote_window_boundaries(self):
        self.assertFalse(is_vote_date("2026-08-10"))
        self.assertTrue(is_vote_date("2026-08-11"))
        self.assertTrue(is_vote_date("2026-08-31 23:59:59"))
        self.assertFalse(is_vote_date("2026-09-01"))

    def test_full_groups_vote_only_after_sign_step(self):
        self.assertFalse(can_vote_after_sign(False))
        self.assertTrue(can_vote_after_sign(False, sign_step_completed=True))
        self.assertTrue(can_vote_after_sign(True))
        self.assertTrue(can_vote_after_sign(False, sign_skipped=True))
        self.assertTrue(
            can_vote_after_sign(
                False,
                data_only_retry=True,
                previous_sign_success=True,
            )
        )
        self.assertFalse(
            can_vote_after_sign(
                False,
                data_only_retry=True,
                previous_sign_success=False,
            )
        )

    def test_vote_ready_already_and_conflict(self):
        target_sku = "target"
        target_name = "product"

        def response(can_vote, voted=""):
            return {
                "success": True,
                "data": {
                    "myVotedProductSku": voted,
                    "voteProductConfigList": [
                        {"productSku": target_sku, "productName": target_name, "canVote": can_vote}
                    ],
                },
            }

        self.assertEqual(inspect_vote_config(response(True), target_sku, target_name)["state"], "ready")
        self.assertEqual(inspect_vote_config(response(False, target_sku), target_sku, target_name)["state"], "already")
        self.assertEqual(inspect_vote_config(response(False, "other"), target_sku, target_name)["state"], "conflict")

    def test_new_winning_response_is_never_truncated(self):
        for count in (0, 1, 8, 15):
            with self.subTest(count=count):
                response = {
                    "success": True,
                    "data": {
                        "records": [
                            {
                                "prizeName": f"prize-{index}",
                                "receiveStatus": 2 if index % 2 else 0,
                            }
                            for index in range(count)
                        ]
                    },
                }
                rows = parse_lottery_winning_response(response)
                self.assertEqual(len(rows), count)
                if rows:
                    self.assertIn("claimed", rows[0])


class DynamicGroupTests(unittest.TestCase):
    def test_schedule_guard_blocks_only_when_manual_run_is_same_shanghai_date(self):
        payload = {
            "workflow_runs": [
                {"event": "workflow_dispatch", "created_at": "2026-08-27T23:30:00Z"},
                {"event": "schedule", "created_at": "2026-08-28T00:00:00Z"},
            ]
        }
        now = datetime.fromisoformat("2026-08-28T07:00:00+08:00")
        self.assertTrue(has_manual_run_today(payload, now))
        tomorrow = datetime.fromisoformat("2026-08-29T07:00:00+08:00")
        self.assertFalse(has_manual_run_today(payload, tomorrow))

    def test_chain_state_base64_round_trip(self):
        state = {
            "schema_version": 1,
            "orchestration_id": "123",
            "task_start_date": "2026-08-26",
            "ref": "main",
            "groups": [{"group_code": "ll3", "account_category": "同行不签到组", "account_count": 234}],
        }
        encoded = encode_chain_state(state)
        self.assertNotIn("{", encoded)
        self.assertEqual(load_chain_state(encoded), state)

    def test_start_chain_dispatches_only_first_group_or_summary_when_empty(self):
        values = {
            f"{prefix}{index}": ""
            for prefix in ("old", "new", "ll", "zh")
            for index in range(1, 21)
        }
        values.update({"old2": "a,p", "new1": "b,p"})
        from h3.dynamic_groups import start_chain

        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, values, clear=False), patch(
            "h3.dynamic_groups.dispatch_group"
        ) as first_group, patch("h3.dynamic_groups.dispatch_summary") as summary:
            args = SimpleNamespace(
                orchestration_id="123",
                ref="main",
                task_start_date="2026-08-26",
                output=os.path.join(temp_dir, "state.json"),
            )
            self.assertEqual(start_chain(args), 0)
            first_group.assert_called_once()
            self.assertEqual(first_group.call_args.args[1], "old2")
            summary.assert_not_called()

        empty_values = {
            f"{prefix}{index}": ""
            for prefix in ("old", "new", "ll", "zh")
            for index in range(1, 21)
        }
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(os.environ, empty_values, clear=False), patch(
            "h3.dynamic_groups.dispatch_group"
        ) as first_group, patch("h3.dynamic_groups.dispatch_summary") as summary:
            args.output = os.path.join(temp_dir, "state.json")
            self.assertEqual(start_chain(args), 0)
            first_group.assert_not_called()
            summary.assert_called_once()

    def test_eighty_group_chain_state_stays_within_dispatch_input_limit(self):
        values = {
            f"{prefix}{index}": "account,password"
            for prefix in ("old", "new", "ll", "zh")
            for index in range(1, 21)
        }
        with patch.dict(os.environ, values, clear=False):
            state = new_chain_state("123", "main", "2026-08-26")
        self.assertEqual(len(state["groups"]), 80)
        self.assertLess(len(compact_json(state).encode("utf-8")), 65535)

    def test_existing_run_prevents_duplicate_dispatch(self):
        existing = {"id": 987, "display_title": "group-123-old1"}
        with patch("h3.dynamic_groups.existing_run", return_value=existing), patch(
            "h3.dynamic_groups.api_request"
        ) as api:
            result = dispatch_once(
                "owner/repo",
                "token",
                "main",
                "dynamic-group.yml",
                "group-123-old1",
                {"group_code": "old1"},
            )
        self.assertEqual(result, {"status": "existing", "run_id": 987})
        api.assert_not_called()

    def test_chain_freezes_configured_groups_and_advances_once(self):
        values = {
            f"{prefix}{index}": ""
            for prefix in ("old", "new", "ll", "zh")
            for index in range(1, 21)
        }
        values.update({"old2": "a,p", "new1": "b,p", "zh3": "c,p"})
        with patch.dict(os.environ, values, clear=False):
            state = new_chain_state("123", "main", "2026-08-26")
        self.assertEqual([item["group_code"] for item in state["groups"]], ["old2", "new1", "zh3"])
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "h3.dynamic_groups.dispatch_group"
        ) as next_group, patch("h3.dynamic_groups.dispatch_summary") as summary:
            output = os.path.join(temp_dir, "state.json")
            args = SimpleNamespace(
                chain_state=json.dumps(state, ensure_ascii=False),
                chain_state_env="",
                current_group="old2",
                current_run_id=456,
                current_result="",
                output=output,
            )
            from h3.dynamic_groups import advance_chain

            self.assertEqual(advance_chain(args), 0)
            next_group.assert_called_once()
            self.assertEqual(next_group.call_args.args[1], "new1")
            summary.assert_not_called()

    def test_chain_last_group_dispatches_summary_and_preserves_run_ids(self):
        state = {
            "schema_version": 1,
            "orchestration_id": "123",
            "task_start_date": "2026-08-26",
            "ref": "main",
            "groups": [
                {"group_code": "old1", "account_category": "老号全干组", "account_count": 1, "run_id": 111},
                {"group_code": "new1", "account_category": "新号全干组", "account_count": 1, "run_id": 0},
            ],
        }
        with tempfile.TemporaryDirectory() as temp_dir, patch("h3.dynamic_groups.dispatch_summary") as summary:
            args = SimpleNamespace(
                chain_state=json.dumps(state, ensure_ascii=False),
                chain_state_env="",
                current_group="new1",
                current_run_id=222,
                current_result="",
                output=os.path.join(temp_dir, "state.json"),
            )
            from h3.dynamic_groups import advance_chain

            self.assertEqual(advance_chain(args), 0)
            summary.assert_called_once()
            updated = json.loads(Path(args.output).read_text(encoding="utf-8"))
            self.assertEqual(updated["groups"][0]["run_id"], 111)
            self.assertEqual(updated["groups"][1]["run_id"], 222)

    def test_download_uses_only_manifest_run_ids_and_keeps_missing_artifact(self):
        archive = io.BytesIO()
        with zipfile.ZipFile(archive, "w") as zipped:
            zipped.writestr("result.json", json.dumps({"group_code": "old1", "results": []}))
        responses = {
            "/actions/runs/111": {"status": "completed", "conclusion": "success", "html_url": "run-url"},
            "/actions/runs/111/artifacts": {"artifacts": [{"id": 9, "name": "group-result-123-old1", "expired": False}]},
            "/actions/runs/222": {"status": "completed", "conclusion": "failure", "html_url": "run-url-2"},
            "/actions/runs/222/artifacts": {"artifacts": []},
        }
        state = {
            "schema_version": 1,
            "orchestration_id": "123",
            "task_start_date": "2026-08-26",
            "ref": "main",
            "groups": [
                {"group_code": "old1", "account_category": "老号全干组", "account_count": 1, "run_id": 111},
                {"group_code": "new1", "account_category": "新号全干组", "account_count": 1, "run_id": 222},
            ],
        }

        def fake_api(method, repo, token, path, **kwargs):
            if path == "/actions/artifacts/9/zip":
                return archive.getvalue()
            return responses[path]

        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ,
            {"GITHUB_REPOSITORY": "owner/repo", "GITHUB_TOKEN": "token", "GITHUB_API_URL": "https://api.example"},
            clear=False,
        ), patch("h3.dynamic_groups.api_request", side_effect=fake_api):
            args = SimpleNamespace(
                chain_state=json.dumps(state, ensure_ascii=False),
                chain_state_env="",
                output_dir=temp_dir,
            )
            self.assertEqual(download_groups(args), 0)
            manifest = json.loads(Path(temp_dir, "manifest.json").read_text(encoding="utf-8"))
            self.assertTrue(manifest["groups"][0]["artifact_downloaded"])
            self.assertEqual(manifest["groups"][1]["artifact_error"], "exact group artifact not found")

    def test_account_display_masks_everything_except_last_five(self):
        self.assertEqual(mask_account("1234567890A"), "******7890A")
        self.assertEqual(mask_account("8565A"), "8565A")

    def test_risk_controlled_failure_has_no_account_detail_in_message(self):
        record = normalize_record(
            {
                "account_index": 1,
                "group_code": "test",
                "account_category": "测试组",
                "risk_controlled": True,
                "sign_success": False,
                "sign_status": "签到风控",
                "detail_reason": "签到失败，疑似违反签到规则",
                "points_fetch_success": True,
                "activity_fetch_success": True,
                "data_fetch_completed": True,
                "account_data_required": False,
                "vote_required": False,
            },
            {"task_start_date": "2026-08-26"},
            {("test", 1): "1234567890A"},
        )
        message, _ = build_message([record], {"task_start_date": "2026-08-26"}, 1)
        self.assertIn("总账号数", message)
        self.assertNotIn("7890A", message)
        self.assertNotIn("疑似违反签到规则", message)

    def test_test_group_lookup_is_limited_to_one_account(self):
        with patch.dict(
            os.environ,
            {
                "TEST": "first,password\nsecond,password",
                "TEST_ACCOUNT_LIMIT": "1",
            },
            clear=False,
        ):
            lookup, total = load_account_lookup()
        self.assertEqual(lookup[("test", 1)], "first")
        self.assertNotIn(("test", 2), lookup)
        self.assertGreaterEqual(total, 1)

    def test_empty_frozen_group_list_does_not_read_later_secrets(self):
        with patch.dict(
            os.environ,
            {
                "old1": "late-account,password",
                "REPORT_GROUP_FILTER_ACTIVE": "true",
                "REPORT_GROUP_CODES": "",
                "REPORT_GROUP_LIMITS": "{}",
            },
            clear=False,
        ):
            lookup, total = load_account_lookup()
        self.assertNotIn(("old1", 1), lookup)

    def test_test_report_manifest_uses_merged_result_count(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            results_dir = os.path.join(temp_dir, "results")
            os.makedirs(os.path.join(results_dir, "test"))
            with open(os.path.join(results_dir, "test", "result.json"), "w", encoding="utf-8") as file:
                json.dump(
                    {
                        "task_start_date": "2026-08-25",
                        "results": [{"account_index": 1, "group_code": "test"}],
                    },
                    file,
                )
            manifest = prepare_manifest(Path(results_dir), "123")
        self.assertEqual(manifest["groups"][0]["account_category"], "测试组")
        self.assertEqual(manifest["groups"][0]["account_count"], 1)

    def test_group_detection_keeps_global_order_and_skips_gaps(self):
        values = {
            f"{prefix}{index}": ""
            for prefix in ("old", "new", "ll", "zh")
            for index in range(1, 21)
        }
        values.update(
            {
                "old2": "old,pw",
                "new1": "new,pw\nnew2,pw",
                "ll3": "peer,pw",
                "zh1": "peer2,pw",
            }
        )
        with patch.dict(os.environ, values, clear=False):
            groups = configured_groups()
        self.assertEqual(
            [item["group_code"] for item in groups],
            ["old2", "new1", "ll3", "zh1"],
        )
        self.assertEqual([item["account_count"] for item in groups], [1, 2, 1, 1])
        self.assertEqual(groups[-1]["account_category"], "同行不签到组")

    def test_skip_sign_is_normal_only_when_data_is_complete(self):
        payload = {"group_code": "ll1", "account_category": "同行不签到组", "execution_mode": "skip_sign"}
        base = {
            "account_index": 1,
            "sign_skipped": True,
            "data_fetch_completed": True,
            "account_data_required": False,
            "points_fetch_success": True,
            "activity_fetch_success": True,
            "listing_gift_required": False,
            "vote_required": False,
        }
        complete = normalize_record(base, payload, {})
        self.assertFalse(is_problem_record(complete))
        incomplete = normalize_record({**base, "data_fetch_completed": False, "points_fetch_success": False}, payload, {})
        self.assertTrue(is_problem_record(incomplete))

    def test_configured_report_path_is_not_rewritten(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.dict(
            os.environ, {"OUTPUT_XLSX_PATH": os.path.join(temp_dir, "2026-08-25-老号全干组.xlsx")}
        ):
            path = resolve_output_xlsx_path(temp_dir, {"task_start_date": "2026-08-25"})
        self.assertTrue(path.endswith("2026-08-25-老号全干组.xlsx"))

    def test_three_category_reports_are_generated_separately(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            results_dir = os.path.join(temp_dir, "results")
            os.makedirs(os.path.join(results_dir, "old1"))
            with open(os.path.join(results_dir, "manifest.json"), "w", encoding="utf-8") as file:
                json.dump(
                    {
                        "task_start_date": "2026-08-25",
                        "groups": [
                            {"group_code": "old1", "account_category": "老号全干组", "account_count": 1}
                        ],
                    },
                    file,
                    ensure_ascii=False,
                )
            with open(os.path.join(results_dir, "old1", "result.json"), "w", encoding="utf-8") as file:
                json.dump(
                    {
                        "group_code": "old1",
                        "account_category": "老号全干组",
                        "results": [
                            {
                                "account_index": 1,
                                "group_code": "old1",
                                "account_category": "老号全干组",
                                "sign_success": True,
                                "sign_status": "签到成功",
                                "points_fetch_success": True,
                                "activity_fetch_success": True,
                                "data_fetch_completed": True,
                                "account_data_required": False,
                                "activity_records": {"lottery": [], "exchange": []},
                                "listing_gift_required": False,
                                "vote_required": False,
                            }
                        ],
                    },
                    file,
                    ensure_ascii=False,
                )
            values = {
                f"{prefix}{index}": ""
                for prefix in ("old", "new", "ll", "zh")
                for index in range(1, 21)
            }
            values.update(
                {
                    "old1": "fixture-account,fixture-password",
                    "NOTIFY_CHANNELS": "",
                    "TELEGRAM_BOT_TOKEN": "",
                    "TELEGRAM_CHAT_ID": "",
                }
            )
            with patch.dict(os.environ, values, clear=False), patch.object(
                sys, "argv", ["category_reports.py", results_dir]
            ):
                self.assertEqual(category_reports_main(), 0)
            names = sorted(name for name in os.listdir(results_dir) if name.endswith(".xlsx"))
            self.assertEqual(
                names,
                [
                    "2026-08-25-同行不签到组.xlsx",
                    "2026-08-25-新号全干组.xlsx",
                    "2026-08-25-老号全干组.xlsx",
                ],
            )

    def test_report_uses_plain_account_and_confidential_password_marker(self):
        values = {
            f"{prefix}{index}": ""
            for prefix in ("old", "new", "ll", "zh")
            for index in range(1, 21)
        }
        values.update({
            "old1": "plain-account,secret-value\nother-account,visible-value",
            "CONFIDENTIAL_PASSWORDS": "secret-value\nsecond-secret",
        })
        with patch.dict(os.environ, values, clear=False):
            lookup, total = load_credential_lookup()
            self.assertEqual(total, 2)
            self.assertEqual(lookup[("old1", 1)]["username"], "plain-account")
            self.assertEqual(report_password(lookup[("old1", 1)]["password"]), "保密")
            self.assertEqual(report_password(lookup[("old1", 2)]["password"]), "visible-value")
            normalized = normalize_record(
                {"account_index": 1, "group_code": "old1"},
                {},
                {("old1", 1): "plain-account"},
            )
            self.assertEqual(normalized["username"], "plain-account")

    def test_telegram_message_is_plain_but_log_message_is_masked(self):
        record = normalize_record(
            {
                "account_index": 1,
                "group_code": "test",
                "sign_success": False,
                "sign_status": "执行异常",
                "detail_reason": "网络超时",
                "account_data_required": False,
                "vote_required": False,
            },
            {},
            {("test", 1): "1234567890A"},
        )
        message, _ = build_message([record], {}, 1)
        self.assertIn("1234567890A", message)
        logged = redact_accounts_for_log(message, [record])
        self.assertNotIn("1234567890A", logged)
        self.assertIn("******7890A", logged)

    def test_retry_matrix_only_contains_unfinished_components(self):
        complete_vote = {
            "account_index": 1,
            "sign_success": False,
            "points_fetch_success": True,
            "account_data": {
                "invoice_fetch_success": True,
                "pcb_order_fetch_success": True,
                "coupon_fetch_success": True,
            },
            "activity_fetch_success": True,
            "listing_gift_required": False,
            "vote_required": True,
            "vote_success": True,
        }
        self.assertEqual(retry_components(complete_vote), ["sign"])
        vote_failure = {**complete_vote, "sign_success": True, "vote_success": False}
        self.assertEqual(retry_components(vote_failure), ["vote"])
        conflict = {
            **vote_failure,
            "vote_detail": "本期已锁定其他商品 fixture-sku",
        }
        self.assertEqual(retry_components(conflict), [])
        password_error = {**vote_failure, "password_error": True}
        self.assertEqual(retry_components(password_error), [])
        partial_activity = {
            **complete_vote,
            "sign_success": True,
            "activity_fetch_success": False,
            "component_status": {"lottery": True, "exchange": False},
        }
        self.assertEqual(retry_components(partial_activity), ["exchange"])

        with tempfile.TemporaryDirectory() as temp_dir:
            result_dir = Path(temp_dir, "initial-result-1")
            result_dir.mkdir()
            Path(result_dir, "result.json").write_text(
                json.dumps({"results": [vote_failure]}, ensure_ascii=False),
                encoding="utf-8",
            )
            matrix = build_retry_matrix(temp_dir, 2)
            self.assertEqual(matrix[0]["account_index"], 1)
            self.assertEqual(matrix[0]["retry_components"], "vote")
            self.assertEqual(matrix[1]["account_index"], 2)

    def test_component_merge_preserves_successful_vote_and_data(self):
        initial = {
            "account_index": 1,
            "sign_success": False,
            "points_fetch_success": True,
            "initial_points": 10,
            "final_points": 10,
            "account_data_fetch_success": True,
            "account_data": {
                "invoice_fetch_success": True,
                "pcb_order_fetch_success": True,
                "coupon_fetch_success": True,
                "coupons": {"unused": [{"name": "fixture"}], "used": [], "expired": []},
            },
            "activity_fetch_success": True,
            "activity_records": {"seckill": [], "lottery": [{"title": "reward"}], "exchange": []},
            "listing_gift_required": False,
            "vote_required": True,
            "vote_success": True,
            "vote_status": "投票成功",
            "retry_count": 0,
        }
        retry = {
            **initial,
            "sign_success": True,
            "account_data_fetch_success": False,
            "account_data": {},
            "activity_fetch_success": False,
            "activity_records": {"seckill": [], "lottery": [], "exchange": []},
            "vote_success": False,
            "vote_status": "未执行",
            "retry_count": 1,
        }
        merged = pick_result(initial, retry)
        self.assertTrue(merged["sign_success"])
        self.assertTrue(merged["vote_success"])
        self.assertEqual(merged["activity_records"]["lottery"][0]["title"], "reward")
        self.assertEqual(merged["account_data"]["coupons"]["unused"][0]["name"], "fixture")

    def test_zero_runner_recovery_only_allows_first_attempt(self):
        run = {"run_attempt": 1, "status": "completed", "conclusion": "failure"}
        jobs = [{"runner_id": 0, "steps": []}, {"runner_id": 0, "steps": []}]
        self.assertTrue(should_rerun(run, jobs))
        self.assertFalse(should_rerun({**run, "run_attempt": 2}, jobs))
        self.assertFalse(should_rerun({**run, "conclusion": "cancelled"}, jobs))
        self.assertFalse(should_rerun(run, [{"runner_id": 7, "steps": []}]))
        self.assertFalse(
            should_rerun(run, [{"runner_id": 0, "steps": [{"status": "completed", "conclusion": "success"}]}])
        )


if __name__ == "__main__":
    unittest.main()
