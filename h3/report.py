import glob
import json
import os
import re
import smtplib
import ssl
import sys
import time
from datetime import datetime
from email.header import Header
from email.mime.text import MIMEText

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from feature_flags import SECKILL_ENABLED
except ImportError:
    from h3.feature_flags import SECKILL_ENABLED

try:
    from account_data import empty_account_data
    from campaign_vote import is_vote_date
    from listing_gift import is_listing_gift_date
except ImportError:
    from h3.account_data import empty_account_data
    from h3.campaign_vote import is_vote_date
    from h3.listing_gift import is_listing_gift_date

for stream_name in ("stdout", "stderr"):
    stream = getattr(sys, stream_name, None)
    if hasattr(stream, "reconfigure"):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass


os.environ.setdefault("TZ", "Asia/Shanghai")
try:
    time.tzset()
except Exception:
    pass

RISK_CONTROL_MESSAGE = (os.getenv("RISK_CONTROL_MESSAGE") or "签到失败，疑似违反签到规则").strip()

STATUS_RED_FILL = PatternFill("solid", fgColor="F8696B")
STATUS_YELLOW_FILL = PatternFill("solid", fgColor="FFD966")
STATUS_BLUE_FILL = PatternFill("solid", fgColor="9DC3E6")
STATUS_GREEN_FILL = PatternFill("solid", fgColor="C6E0B4")
PRIZE_ALERT_FILL = PatternFill("solid", fgColor="F8696B")
FONT_GREEN = Font(color="008000")
FONT_RED = Font(color="C00000")
FONT_BLUE = Font(color="1F4E79")
FONT_DARK = Font(color="000000")


def truthy(value) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def safe_int(value, default=0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def safe_float(value, default=0.0) -> float:
    try:
        return float(str(value).strip())
    except Exception:
        return default


def date_part(value="") -> str:
    match = re.search(r"\d{4}-\d{2}-\d{2}", str(value or ""))
    return match.group(0) if match else ""


def default_group_name(group_number: int) -> str:
    return f"{group_number}组" if group_number > 0 else ""


def default_group_position(group_number: int, account_index: int) -> str:
    if group_number > 0 and account_index > 0:
        return f"{group_number}组账号{account_index}"
    return f"账号{account_index}" if account_index > 0 else "未知账号"


def load_account_lookup() -> tuple[dict[tuple[object, int], str], int]:
    lookup = {}
    total = 0
    for prefix in ("old", "new", "ll", "zh"):
        for slot in range(1, 21):
            group_code = f"{prefix}{slot}"
            raw = os.getenv(group_code) or ""
            for account_index, line in enumerate(
                (line.strip() for line in raw.splitlines() if line.strip() and "," in line),
                start=1,
            ):
                lookup[(group_code, account_index)] = line.split(",", 1)[0].strip()
                total += 1
    raw_test = os.getenv("test") or os.getenv("TEST") or ""
    for account_index, line in enumerate(
        (line.strip() for line in raw_test.splitlines() if line.strip() and "," in line),
        start=1,
    ):
        lookup[("test", account_index)] = line.split(",", 1)[0].strip()
        total += 1
    for group_number in range(1, 9):
        raw = os.getenv(f"ACCOUNTS_BATCH{group_number}", "") or ""
        for account_index, line in enumerate(raw.splitlines(), start=1):
            line = line.strip()
            if not line or "," not in line:
                continue
            username = line.split(",", 1)[0].strip()
            lookup[(group_number, account_index)] = username
            total += 1
    return lookup, total


def load_manifest(results_dir: str) -> dict:
    path = os.path.join(results_dir, "manifest.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return {}


def target_date_text(manifest: dict) -> str:
    if isinstance(manifest, dict) and manifest.get("target_date"):
        return str(manifest["target_date"]).strip()
    if isinstance(manifest, dict) and manifest.get("task_start_date"):
        return date_part(manifest["task_start_date"]) or str(manifest["task_start_date"]).strip()
    return datetime.now().strftime("%Y-%m-%d")


def resolve_output_xlsx_path(results_dir: str, manifest: dict) -> str:
    filename = f"{target_date_text(manifest)}签到汇总.xlsx"
    configured_path = os.getenv("OUTPUT_XLSX_PATH") or ""
    if configured_path:
        return configured_path
    return os.path.join(results_dir, filename)


def find_json_files(results_dir: str) -> list[str]:
    paths = set()
    for path in glob.glob(os.path.join(results_dir, "**", "*.json"), recursive=True):
        if os.path.basename(path).lower() == "manifest.json":
            continue
        if os.path.isfile(path):
            paths.add(path)
    return sorted(paths)


def record_key(record: dict):
    group_code = str(record.get("group_code") or "").strip().lower()
    group_number = safe_int(record.get("group_number"), 0)
    account_index = safe_int(record.get("account_index"), 0)
    if group_code and account_index > 0:
        return group_code, account_index
    if group_number > 0 and account_index > 0:
        return group_number, account_index
    return None


def normalize_activity_records(value) -> dict:
    if not isinstance(value, dict):
        return {"seckill": [], "lottery": [], "exchange": []}

    normalized = {"seckill": [], "lottery": [], "exchange": []}
    activity_types = (("seckill", 2), ("lottery", None)) if SECKILL_ENABLED else (("lottery", None),)
    for key, limit in activity_types:
        rows = value.get(key)
        if not isinstance(rows, list):
            continue
        selected_rows = rows if limit is None else rows[:limit]
        for item in selected_rows:
            if not isinstance(item, dict):
                continue
            normalized[key].append(
                {
                    "title": str(item.get("title") or item.get("skuTitle") or item.get("prizeTitle") or "").strip(),
                    "status_text": str(item.get("status_text") or "").strip(),
                    "claimed": truthy(item.get("claimed")),
                    "expiry_date": str(item.get("expiry_date") or "").strip(),
                }
            )
    exchange_rows = value.get("exchange")
    if isinstance(exchange_rows, list):
        for item in exchange_rows:
            if not isinstance(item, dict):
                continue
            title = str(item.get("title") or item.get("goodsName") or "").strip()
            if not title:
                continue
            normalized["exchange"].append(
                {
                    "title": title,
                    "status": safe_int(item.get("status"), None),
                    "status_text": str(item.get("status_text") or "").strip(),
                    "quantity": max(1, safe_int(item.get("quantity"), 1)),
                    "points": safe_float(item.get("points"), 0.0),
                    "created_at": str(item.get("created_at") or "").strip(),
                }
            )
    return normalized


def normalize_record(record: dict, payload: dict, account_lookup: dict[tuple[object, int], str]) -> dict:
    group_number = safe_int(record.get("group_number", payload.get("group_number")), 0)
    account_index = safe_int(record.get("account_index"), 0)
    group_code = str(record.get("group_code") or payload.get("group_code") or "").strip().lower()
    username = str(
        record.get("username")
        or record.get("masked_username")
        or account_lookup.get((group_code, account_index))
        or account_lookup.get((group_number, account_index))
        or f"账号{account_index}"
    ).strip()
    detail_reason = str(record.get("detail_reason") or "").strip()
    risk_controlled = truthy(record.get("risk_controlled")) or (RISK_CONTROL_MESSAGE and RISK_CONTROL_MESSAGE in detail_reason)
    banned_account = truthy(record.get("banned_account"))
    legacy_banned_complete = (
        banned_account
        and "BANNED_ACCOUNTS" in detail_reason
        and not any(text in detail_reason for text in ("获取失败", "未完成", "Token提取失败", "执行异常"))
    )
    points_fetch_success = truthy(record.get("points_fetch_success")) or legacy_banned_complete
    activity_fetch_success = truthy(record.get("activity_fetch_success")) or legacy_banned_complete
    account_data_required = truthy(record.get("account_data_required"))
    account_data_fetch_success = truthy(record.get("account_data_fetch_success"))
    data_fetch_completed = (
        truthy(record.get("data_fetch_completed"))
        or (points_fetch_success and activity_fetch_success)
    ) and (not account_data_required or account_data_fetch_success)
    group_name = str(record.get("group_name") or payload.get("group_name") or payload.get("batch_name") or default_group_name(group_number)).strip()
    group_position = str(record.get("group_position") or default_group_position(group_number, account_index)).strip()
    task_start_date = str(record.get("task_start_date") or payload.get("task_start_date") or "").strip()
    sign_time = str(record.get("sign_time") or record.get("sign_completed_at") or "").strip()
    next_day_success = truthy(record.get("next_day_success")) or (
        truthy(record.get("sign_success")) and date_part(task_start_date) and date_part(sign_time) and date_part(sign_time) > date_part(task_start_date)
    )
    account_category = str(record.get("account_category") or payload.get("account_category") or "").strip()
    execution_mode = str(record.get("execution_mode") or payload.get("execution_mode") or "").strip()
    sign_skipped = truthy(record.get("sign_skipped")) or execution_mode == "skip_sign"
    vote_required = truthy(record.get("vote_required"))
    vote_success = truthy(record.get("vote_success"))
    vote_status = str(record.get("vote_status") or ("待执行" if vote_required else "非投票日期")).strip()
    listing_gift_required = (
        truthy(record.get("listing_gift_required"))
        if "listing_gift_required" in record
        else is_listing_gift_date(task_start_date)
    )
    listing_gift_status = str(record.get("listing_gift_status") or "").strip()
    if listing_gift_required and not listing_gift_status:
        listing_gift_status = "缺少礼包领取结果"
    return {
        "account_index": account_index,
        "execution_order": safe_int(record.get("execution_order"), 0),
        "username": username,
        "group_name": group_name,
        "group_number": group_number,
        "group_position": group_position,
        "group_code": group_code,
        "account_category": account_category,
        "execution_mode": execution_mode,
        "sign_skipped": sign_skipped,
        "sign_success": truthy(record.get("sign_success")),
        "sign_status": str(record.get("sign_status") or "").strip(),
        "initial_points": safe_float(record.get("initial_points"), 0.0),
        "final_points": safe_float(record.get("final_points"), 0.0),
        "points_reward": safe_float(record.get("points_reward"), 0.0),
        "has_reward": truthy(record.get("has_reward")),
        "password_error": truthy(record.get("password_error")),
        "risk_controlled": risk_controlled,
        "banned_account": banned_account,
        "points_fetch_success": points_fetch_success,
        "activity_fetch_success": activity_fetch_success,
        "data_fetch_completed": data_fetch_completed,
        "next_day_success": next_day_success,
        "task_start_date": task_start_date,
        "sign_completed_at": str(record.get("sign_completed_at") or "").strip(),
        "retry_count": safe_int(record.get("retry_count"), 0),
        "is_final_retry": truthy(record.get("is_final_retry")),
        "detail_reason": detail_reason,
        "sign_time": sign_time,
        "sign_ip": str(record.get("sign_ip") or "").strip(),
        "activity_records": normalize_activity_records(record.get("activity_records")),
        "account_data_required": account_data_required,
        "account_data_fetch_success": account_data_fetch_success,
        "account_data": record.get("account_data") if isinstance(record.get("account_data"), dict) else empty_account_data(),
        "listing_gift_required": listing_gift_required,
        "listing_gift_success": truthy(record.get("listing_gift_success")),
        "listing_gift_attempted": truthy(record.get("listing_gift_attempted")),
        "listing_gift_status": listing_gift_status,
        "listing_gift_time": str(record.get("listing_gift_time") or "").strip(),
        "listing_gift_detail": str(record.get("listing_gift_detail") or "").strip(),
        "vote_required": vote_required,
        "vote_success": vote_success,
        "vote_attempted": truthy(record.get("vote_attempted")),
        "vote_status": vote_status,
        "vote_time": str(record.get("vote_time") or "").strip(),
        "vote_product_sku": str(record.get("vote_product_sku") or "").strip(),
        "vote_product_name": str(record.get("vote_product_name") or "").strip(),
        "vote_detail": str(record.get("vote_detail") or "").strip(),
    }


def load_results(results_dir: str, account_lookup: dict[tuple[object, int], str]) -> list[dict]:
    records_by_key = {}
    extras = []
    for path in find_json_files(results_dir):
        try:
            with open(path, "r", encoding="utf-8") as file:
                payload = json.load(file)
        except Exception:
            continue
        rows = payload.get("results") if isinstance(payload, dict) else None
        if not isinstance(rows, list):
            continue
        for record in rows:
            if not isinstance(record, dict):
                continue
            normalized = normalize_record(record, payload, account_lookup)
            key = record_key(normalized)
            if key is None:
                extras.append(normalized)
            else:
                records_by_key[key] = normalized
    return list(records_by_key.values()) + extras


def build_missing_record(group_identity, account_index: int, username: str, task_date: str = "") -> dict:
    group_code = str(group_identity).lower() if isinstance(group_identity, str) else ""
    group_number = safe_int(group_identity, 0) if not group_code else 0
    if group_code.startswith("old"):
        category = "老号全干组"
    elif group_code.startswith("new"):
        category = "新号全干组"
    elif group_code.startswith(("ll", "zh")):
        category = "同行不签到组"
    elif group_code == "test":
        category = "测试组"
    else:
        category = ""
    sign_skipped = group_code.startswith(("ll", "zh"))
    vote_required = is_vote_date(task_date)
    return {
        "account_index": account_index,
        "execution_order": account_index,
        "username": username,
        "group_name": default_group_name(group_number),
        "group_number": group_number,
        "group_position": f"{group_code}账号{account_index}" if group_code else default_group_position(group_number, account_index),
        "group_code": group_code,
        "account_category": category,
        "execution_mode": "skip_sign" if sign_skipped else "full",
        "sign_skipped": sign_skipped,
        "sign_success": False,
        "sign_status": "签到异常",
        "initial_points": 0.0,
        "final_points": 0.0,
        "points_reward": 0.0,
        "has_reward": False,
        "password_error": False,
        "risk_controlled": False,
        "banned_account": False,
        "points_fetch_success": False,
        "activity_fetch_success": False,
        "data_fetch_completed": False,
        "next_day_success": False,
        "task_start_date": task_date,
        "sign_completed_at": "",
        "retry_count": 0,
        "is_final_retry": False,
        "detail_reason": "缺少签到结果",
        "sign_time": "",
        "sign_ip": "",
        "activity_records": {"seckill": [], "lottery": [], "exchange": []},
        "account_data_required": truthy(os.getenv("ACCOUNT_DATA_ENABLED", "false")),
        "account_data_fetch_success": False,
        "account_data": empty_account_data(),
        "listing_gift_required": is_listing_gift_date(task_date),
        "listing_gift_success": False,
        "listing_gift_attempted": False,
        "listing_gift_status": "缺少礼包领取结果" if is_listing_gift_date(task_date) else "非领取日期",
        "listing_gift_time": "",
        "listing_gift_detail": "",
        "vote_required": vote_required,
        "vote_success": False,
        "vote_attempted": False,
        "vote_status": "缺少投票结果" if vote_required else "非投票日期",
        "vote_time": "",
        "vote_product_sku": "",
        "vote_product_name": "",
        "vote_detail": "",
    }


def merge_records_with_expected(
    records: list[dict],
    account_lookup: dict[tuple[object, int], str],
    target_date: str = "",
) -> list[dict]:
    indexed = {}
    extras = []
    for record in records:
        key = record_key(record)
        if key is None:
            extras.append(record)
        else:
            indexed[key] = record

    if not account_lookup:
        return list(indexed.values()) + extras

    merged = []
    for key in sorted(account_lookup, key=lambda item: (str(item[0]), item[1])):
        record = indexed.pop(key, None)
        if record is None:
            merged.append(build_missing_record(key[0], key[1], account_lookup[key], target_date))
            continue
        if not record.get("username"):
            record["username"] = account_lookup[key]
        merged.append(record)

    unexpected = sorted(
        list(indexed.values()) + extras,
        key=lambda item: (
            safe_int(item.get("group_number"), 999999),
            safe_int(item.get("account_index"), 999999),
            str(item.get("username") or ""),
        ),
    )
    merged.extend(unexpected)
    return merged


def status_label(record: dict) -> str:
    raw_status = str(record.get("sign_status") or "")
    if truthy(record.get("banned_account")):
        if not truthy(record.get("data_fetch_completed")):
            return "签到异常"
        return "账号封禁"
    if truthy(record.get("sign_skipped")):
        return "按配置跳过签到" if truthy(record.get("data_fetch_completed")) else "取数异常"
    if truthy(record.get("next_day_success")):
        return "签到成功但次日"
    if truthy(record.get("risk_controlled")):
        return "签到风控"
    if truthy(record.get("sign_success")):
        return "签到成功"
    if truthy(record.get("password_error")) or any(keyword in raw_status for keyword in ("失败", "错误", "Token", "token")):
        return "签到失败"
    return "签到异常"


def detail_reason(record: dict) -> str:
    reason = str(record.get("detail_reason") or "").strip()
    if reason:
        return reason
    if record.get("risk_controlled"):
        return RISK_CONTROL_MESSAGE
    if record.get("sign_status"):
        return str(record["sign_status"]).strip()
    return "签到异常"


def detail_text(record: dict) -> str:
    if truthy(record.get("banned_account")):
        reason = str(record.get("detail_reason") or "").strip()
        if reason:
            return reason
        status = str(record.get("sign_status") or "").strip()
        return f"账号在封禁列表中，已跳过签到；数据获取失败：{status or '未知异常'}"
    if truthy(record.get("sign_skipped")):
        return str(record.get("detail_reason") or record.get("sign_status") or "同行组按配置跳过签到").strip()
    if truthy(record.get("sign_success")):
        reason = str(record.get("detail_reason") or "").strip()
        activity_failure_markers = (
            "秒杀数据获取失败",
            "抽奖数据获取失败",
            "奖品过期记录获取失败",
            "兑换记录获取失败",
            "活动数据抓取异常",
            "会员资料接口未完整返回",
            "会员中心 SSO 未建立",
        )
        if any(marker in reason for marker in activity_failure_markers):
            status = str(record.get("sign_status") or "签到成功").strip()
            return f"{status}；{reason}" if status not in reason else reason
        return str(record.get("sign_status") or "签到成功").strip()
    return detail_reason(record)


def is_problem_record(record: dict) -> bool:
    return (
        status_sort_bucket(record) == 0
        or not truthy(record.get("data_fetch_completed"))
        or (truthy(record.get("listing_gift_required")) and not truthy(record.get("listing_gift_success")))
        or (truthy(record.get("vote_required")) and not truthy(record.get("vote_success")))
    )


def problem_reason(record: dict) -> str:
    reasons = []
    if status_sort_bucket(record) == 0:
        reasons.append(detail_reason(record))
    if truthy(record.get("listing_gift_required")) and not truthy(record.get("listing_gift_success")):
        reasons.append(str(record.get("listing_gift_status") or record.get("listing_gift_detail") or "礼包领取未完成").strip())
    if not truthy(record.get("data_fetch_completed")):
        reasons.append("账号数据获取未完成")
    if truthy(record.get("vote_required")) and not truthy(record.get("vote_success")):
        reasons.append(str(record.get("vote_status") or record.get("vote_detail") or "投票未完成").strip())
    return "；".join(dict.fromkeys(reason for reason in reasons if reason)) or "未知异常"


def status_sort_bucket(record: dict) -> int:
    label = status_label(record)
    if label in {"签到失败", "签到异常", "签到风控", "取数异常"}:
        return 0
    if label == "签到成功但次日":
        return 1
    return 2


def sort_records(records: list[dict]) -> list[dict]:
    return sorted(
        records,
        key=lambda item: (
            status_sort_bucket(item),
            -safe_float(item.get("final_points"), 0.0),
            safe_int(item.get("group_number"), 999999),
            safe_int(item.get("account_index"), 999999),
            str(item.get("username") or ""),
        ),
    )


def format_percent(value: float) -> str:
    return f"{value:.1f}".rstrip("0").rstrip(".")


def build_summary(records: list[dict], expected_total: int) -> dict:
    total = expected_total or len(records)
    success = sum(1 for item in records if status_label(item) in {"签到成功", "签到成功但次日"})
    banned = sum(1 for item in records if status_label(item) == "账号封禁")
    next_day = sum(1 for item in records if status_label(item) == "签到成功但次日")
    risk = sum(1 for item in records if status_label(item) == "签到风控")
    failed = sum(1 for item in records if status_label(item) == "签到失败")
    abnormal = sum(1 for item in records if status_label(item) == "签到异常")
    skipped = sum(1 for item in records if status_label(item) == "按配置跳过签到")
    reward = sum(safe_float(item.get("points_reward"), 0.0) for item in records)
    success_rate = (success / total * 100) if total > 0 else 0.0
    listing_gift_required = sum(1 for item in records if truthy(item.get("listing_gift_required")))
    listing_gift_success = sum(
        1 for item in records
        if truthy(item.get("listing_gift_required")) and truthy(item.get("listing_gift_success"))
    )
    return {
        "total": total,
        "success": success,
        "banned": banned,
        "next_day": next_day,
        "risk": risk,
        "failed": failed,
        "abnormal": abnormal,
        "skipped": skipped,
        "problem_count": sum(1 for item in records if is_problem_record(item)),
        "reward": reward,
        "success_rate": success_rate,
        "listing_gift_required": listing_gift_required,
        "listing_gift_success": listing_gift_success,
    }


def build_stats_lines(summary: dict) -> list[str]:
    return [
        "📈 总体统计",
        f"  ├── 总账号数: {summary['total']}",
        f"  ├── 签到成功: {summary['success']}/{summary['total']}",
        f"  ├── 次日成功: {summary['next_day']}",
        f"  ├── 按配置跳过签到: {summary['skipped']}",
        f"  ├── 账号封禁: {summary['banned']}",
        f"  ├── 上市礼包完成: {summary['listing_gift_success']}/{summary['listing_gift_required']}",
        f"  ├── 总计获得 +{summary['reward']:.1f} 🌽",
        f"  └── 签到成功率: {format_percent(summary['success_rate'])}%",
    ]


def build_message(records: list[dict], manifest: dict, expected_total: int) -> tuple[str, dict]:
    sorted_records = sort_records(records)
    summary = build_summary(sorted_records, expected_total)
    problem_records = [record for record in sorted_records if is_problem_record(record)]
    category_label = str(os.getenv("SUMMARY_CATEGORY_LABEL") or "").strip()

    if category_label and not sorted_records and expected_total == 0:
        lines = [f"{category_label}：本次未配置账号"]
        lines.extend(build_stats_lines(summary))
        return "\n".join(lines), summary

    if problem_records:
        lines = ["NO❗今天出现问题了捏"]
        for record in problem_records:
            lines.append(f"{record['username']}：{problem_reason(record)}❌")
        lines.extend(build_stats_lines(summary))
        return "\n".join(lines), summary

    if not sorted_records:
        lines = ["NO❗今天出现问题了捏", "未读取到任何签到结果❌"]
        lines.extend(build_stats_lines(summary))
        return "\n".join(lines), summary

    lines = ["喵喵~今天一切正常捏"]
    lines.extend(build_stats_lines(summary))
    return "\n".join(lines), summary


def color_for_points(points: float):
    if points > 2000:
        return PatternFill("solid", fgColor="F8696B")
    if points > 1000:
        return PatternFill("solid", fgColor="FFD966")
    if points > 500:
        return PatternFill("solid", fgColor="9DC3E6")
    if points < 200:
        return PatternFill("solid", fgColor="C6E0B4")
    if 200 <= points < 300:
        return PatternFill("solid", fgColor="DAF2D0")
    if 300 <= points <= 500:
        return PatternFill("solid", fgColor="F4CCCC")
    return None


def font_for_status(label: str) -> Font:
    if label in {"签到失败", "签到异常", "签到风控"}:
        return Font(color="FFFFFF", bold=True)
    if label == "签到成功但次日":
        return Font(color="9C6500", bold=True)
    if label == "账号封禁":
        return Font(color="FFFFFF", bold=True)
    if label in {"签到成功", "按配置跳过签到"}:
        return FONT_GREEN
    return FONT_DARK


def fill_for_status(label: str):
    if label in {"签到失败", "签到异常", "签到风控"}:
        return STATUS_RED_FILL
    if label == "签到成功但次日":
        return STATUS_YELLOW_FILL
    if label == "账号封禁":
        return STATUS_BLUE_FILL
    return None


def font_for_vote_status(record: dict) -> Font:
    if truthy(record.get("vote_success")):
        return FONT_GREEN
    if truthy(record.get("vote_required")):
        return FONT_RED
    return FONT_DARK


def font_for_claim_status(value: str) -> Font:
    text = str(value or "")
    if "已经领取" in text:
        return FONT_GREEN
    if "未领取" in text or "暂未领取" in text or "已过期" in text:
        return FONT_RED
    return FONT_DARK


def fill_for_prize(value: str):
    text = re.sub(r"\s+", "", str(value or ""))
    if not text or text == "6金豆":
        return None
    return PRIZE_ALERT_FILL


def activity_status_text(item: dict) -> str:
    status_text = str(item.get("status_text") or "").strip()
    if status_text:
        return status_text
    if truthy(item.get("claimed")):
        return "已经领取"
    expiry_date = str(item.get("expiry_date") or "").strip()
    return f"未领取 {expiry_date}".strip() if expiry_date else "未领取"


def activity_columns(record: dict) -> list[str]:
    activity = normalize_activity_records(record.get("activity_records"))
    values = []
    activity_types = (("seckill", 2), ("lottery", None)) if SECKILL_ENABLED else (("lottery", None),)
    for key, limit in activity_types:
        rows = activity.get(key) or []
        count = limit if limit is not None else len(rows)
        for index in range(count):
            item = rows[index] if index < len(rows) else {}
            if item:
                values.extend([str(item.get("title") or "").strip(), activity_status_text(item)])
            else:
                values.extend(["", ""])
    return values


def max_lottery_count(records: list[dict]) -> int:
    return max(
        (len((normalize_activity_records(record.get("activity_records")).get("lottery") or [])) for record in records),
        default=0,
    )


def exchange_prize_titles(records: list[dict]) -> list[str]:
    titles = []
    seen = set()
    for record in sort_records(records):
        for item in normalize_activity_records(record.get("activity_records")).get("exchange", []):
            title = str(item.get("title") or "").strip()
            if title and title not in seen:
                seen.add(title)
                titles.append(title)
    return titles


def exchange_detail_text(item: dict) -> str:
    quantity = max(1, safe_int(item.get("quantity"), 1))
    created_at = str(item.get("created_at") or "").strip()
    points = safe_float(item.get("points"), 0.0)
    values = [f"{quantity}件"]
    if created_at:
        values.append(created_at)
    if points:
        values.append(f"{points:g}金豆")
    return "，".join(values)


def exchange_columns(record: dict, prize_titles: list[str]) -> list[str]:
    grouped = {}
    for item in normalize_activity_records(record.get("activity_records")).get("exchange", []):
        grouped.setdefault(str(item.get("title") or "").strip(), []).append(item)
    values = []
    for title in prize_titles:
        rows = grouped.get(title, [])
        values.extend(
            [
                "\n".join(exchange_detail_text(item) for item in rows),
                "\n".join(str(item.get("status_text") or "").strip() for item in rows),
            ]
        )
    return values


def fill_for_exchange_status(value: str):
    statuses = {line.strip() for line in str(value or "").splitlines() if line.strip()}
    if "已退回" in statuses:
        return STATUS_RED_FILL
    if statuses.intersection({"已兑换待发货", "已确认收货"}):
        return STATUS_GREEN_FILL
    if "已兑换" in statuses:
        return STATUS_YELLOW_FILL
    return None


def coupon_summary_text(coupons: list[dict]) -> str:
    values = []
    for coupon in coupons or []:
        detail = str(coupon.get("name") or "未命名优惠券").strip()
        business = str(coupon.get("business_type") or "").strip()
        expiry = str(coupon.get("expires_at") or "").strip()
        rule = str(coupon.get("rule_text") or "").strip()
        if business:
            detail += f"（{business}）"
        if expiry:
            detail += f"，过期：{expiry}"
        if rule:
            detail += f"，规则：{rule}"
        values.append(detail)
    return "\n".join(values)


def invoice_amount_text(value) -> str:
    if value in (None, ""):
        return ""
    number = safe_float(value, None)
    return str(value) if number is None else f"{number:g}"


def safe_sheet_title(raw: str, used: set[str]) -> str:
    title = re.sub(r"[\\/*?:\[\]]", "_", str(raw or "未命名优惠券")).strip() or "未命名优惠券"
    title = title[:31]
    candidate = title
    serial = 2
    while candidate.lower() in {item.lower() for item in used}:
        suffix = f"_{serial}"
        candidate = f"{title[:31-len(suffix)]}{suffix}"
        serial += 1
    used.add(candidate)
    return candidate


def write_coupon_sheets(workbook, records: list[dict]):
    by_name = {}
    for record in sort_records(records):
        account = str(record.get("username") or "").strip()
        for coupon in (record.get("account_data") or {}).get("coupons", {}).get("unused", []) or []:
            name = str(coupon.get("name") or "未命名优惠券").strip()
            by_name.setdefault(name, []).append(
                (account, str(coupon.get("expires_at") or "").strip())
            )
    used = {workbook.active.title}
    for name, rows in sorted(by_name.items(), key=lambda item: item[0]):
        sheet = workbook.create_sheet(safe_sheet_title(name, used))
        sheet.append(["序号", "账户", "优惠券过期时间"])
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="E2F0D9")
            cell.font = Font(bold=True)
        for index, (account, expiry) in enumerate(rows, start=1):
            sheet.append([index, account, expiry])
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 24


def write_xlsx(path: str, records: list[dict]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "签到汇总"
    lottery_count = max_lottery_count(records)
    exchange_titles = exchange_prize_titles(records)
    thresholds = {
        safe_int((record.get("account_data") or {}).get("invoice_month_threshold"), 12) or 12
        for record in records
        if (record.get("account_data") or {}).get("invoice_month_threshold")
    }
    threshold_text = str(next(iter(thresholds))) if len(thresholds) == 1 else "接口阈值"
    headers = [
        "序号",
        "金豆数量",
        "账户",
        "组别",
        "账号类别",
        "执行模式",
        "签到情况",
        "详细原因",
        "签到时间",
        "签到IP",
        "开票资料",
        f"不超过{threshold_text}个月可开金额" if threshold_text != "接口阈值" else "不超过接口月份阈值可开金额",
        f"超过{threshold_text}个月可开金额" if threshold_text != "接口阈值" else "超过接口月份阈值可开金额",
        "PCB 12个月内消费",
        "PCB 超过12个月消费",
        "PCB累计消费",
        "距离40元还差",
        "未使用优惠券",
        "已使用优惠券",
        "已过期优惠券",
        "PCB+SMT优惠券预测",
        "预测依据",
        "上市礼包领取情况",
        "投票状态",
        "投票时间",
        "投票商品",
        "投票详情",
    ]
    for title in exchange_titles:
        headers.extend([f"兑换物品：{title}", f"兑换状态：{title}"])
    activity_start_column = len(headers) + 1
    if SECKILL_ENABLED:
        headers.extend([
            "秒杀一",
            "领取情况",
            "秒杀二",
            "领取情况",
        ])
    for index in range(1, lottery_count + 1):
        headers.extend([f"抽奖{index}", f"领取情况{index}"])
    sheet.append(headers)
    header_index = {name: index for index, name in enumerate(headers, start=1)}

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for index, record in enumerate(sort_records(records), start=1):
        label = status_label(record)
        row = [
            index,
            safe_float(record.get("final_points"), 0.0),
            str(record.get("username") or ""),
            str(record.get("group_code") or record.get("group_position") or ""),
            str(record.get("account_category") or ""),
            str(record.get("execution_mode") or ""),
            label,
            detail_text(record),
            str(record.get("sign_time") or ""),
            str(record.get("sign_ip") or ""),
            str((record.get("account_data") or {}).get("invoice_profile_status") or "数据不足"),
            invoice_amount_text((record.get("account_data") or {}).get("invoice_within_months_amount")),
            invoice_amount_text((record.get("account_data") or {}).get("invoice_over_months_amount")),
            invoice_amount_text((record.get("account_data") or {}).get("pcb_within_months_amount")),
            invoice_amount_text((record.get("account_data") or {}).get("pcb_over_months_amount")),
            invoice_amount_text((record.get("account_data") or {}).get("pcb_total_amount")),
            invoice_amount_text((record.get("account_data") or {}).get("pcb_amount_shortfall")),
            coupon_summary_text((record.get("account_data") or {}).get("coupons", {}).get("unused")),
            coupon_summary_text((record.get("account_data") or {}).get("coupons", {}).get("used")),
            coupon_summary_text((record.get("account_data") or {}).get("coupons", {}).get("expired")),
            str((record.get("account_data") or {}).get("coupon_prediction") or "数据不足"),
            str((record.get("account_data") or {}).get("prediction_reason") or ""),
            "\n".join(
                value
                for value in (
                    str(record.get("listing_gift_status") or "").strip(),
                    str(record.get("listing_gift_time") or "").strip(),
                )
                if value
            ),
            str(record.get("vote_status") or ""),
            str(record.get("vote_time") or ""),
            str(record.get("vote_product_name") or record.get("vote_product_sku") or ""),
            str(record.get("vote_detail") or ""),
        ] + exchange_columns(record, exchange_titles) + activity_columns(record)
        sheet.append(row)
        row_index = sheet.max_row
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        sheet.cell(row_index, 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row_index, 2).alignment = Alignment(horizontal="center", vertical="center")
        for name in ("组别", "账号类别", "执行模式", "签到情况", "签到时间", "签到IP"):
            sheet.cell(row_index, header_index[name]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row_index, header_index["详细原因"]).alignment = Alignment(vertical="center", wrap_text=True)
        for column_index in range(header_index["开票资料"], len(headers) + 1):
            sheet.cell(row_index, column_index).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row_index, 2).number_format = "0.0"
        fill = color_for_points(safe_float(record.get("final_points"), 0.0))
        if fill:
            sheet.cell(row_index, 2).fill = fill
        status_fill = fill_for_status(label)
        if status_fill:
            sheet.cell(row_index, header_index["签到情况"]).fill = status_fill
        sheet.cell(row_index, header_index["签到情况"]).font = font_for_status(label)
        invoice_cell = sheet.cell(row_index, header_index["开票资料"])
        invoice_cell.font = FONT_GREEN if invoice_cell.value == "有" else (FONT_RED if invoice_cell.value == "无" else FONT_BLUE)
        prediction_cell = sheet.cell(row_index, header_index["PCB+SMT优惠券预测"])
        prediction_cell.font = FONT_RED if prediction_cell.value in {"不可能", "很小可能"} else (FONT_GREEN if prediction_cell.value in {"很大可能", "100%可能"} else FONT_BLUE)
        gift_cell = sheet.cell(row_index, header_index["上市礼包领取情况"])
        gift_cell.font = FONT_GREEN if truthy(record.get("listing_gift_success")) else (FONT_RED if truthy(record.get("listing_gift_required")) else FONT_DARK)
        vote_cell = sheet.cell(row_index, header_index["投票状态"])
        vote_cell.font = font_for_vote_status(record)
        for title in exchange_titles:
            status_column = header_index[f"兑换状态：{title}"]
            exchange_fill = fill_for_exchange_status(sheet.cell(row_index, status_column).value)
            if exchange_fill:
                sheet.cell(row_index, status_column).fill = exchange_fill
        for column_index in range(activity_start_column, len(headers) + 1, 2):
            prize_fill = fill_for_prize(sheet.cell(row_index, column_index).value)
            if prize_fill:
                sheet.cell(row_index, column_index).fill = prize_fill
        for column_index in range(activity_start_column + 1, len(headers) + 1, 2):
            sheet.cell(row_index, column_index).font = font_for_claim_status(sheet.cell(row_index, column_index).value)

    sheet.freeze_panes = "A2"
    widths = {}
    for name, column_index in header_index.items():
        width = 18
        if name == "序号":
            width = 8
        elif name in {"账户", "投票商品"}:
            width = 24
        elif name in {"详细原因", "预测依据", "投票详情"}:
            width = 42
        elif name in {"未使用优惠券", "已使用优惠券", "已过期优惠券"}:
            width = 36
        elif name.startswith("兑换物品："):
            width = 30
        elif name.startswith("兑换状态："):
            width = 20
        widths[get_column_letter(column_index)] = width
    for column_index in range(activity_start_column, len(headers) + 1):
        widths[get_column_letter(column_index)] = 28 if (column_index - activity_start_column) % 2 == 0 else 18
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    write_coupon_sheets(workbook, records)
    workbook.save(path)


def split_text(text: str, limit: int = 3900) -> list[str]:
    if len(text) <= limit:
        return [text]
    parts = []
    current = ""
    for line in text.splitlines(True):
        if len(current) + len(line) > limit and current:
            parts.append(current)
            current = ""
        current += line
    if current:
        parts.append(current)
    return parts


def send_telegram_message(text: str) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        return False
    ok = True
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    for part in split_text(text):
        try:
            response = requests.post(url, json={"chat_id": chat_id, "text": part}, timeout=20)
            if response.status_code != 200:
                ok = False
        except Exception:
            ok = False
    return ok


def send_telegram_document(path: str) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id or not os.path.exists(path):
        return False
    try:
        with open(path, "rb") as file:
            response = requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={"document": (os.path.basename(path), file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=40,
            )
        return response.status_code == 200
    except Exception:
        return False


def send_email(subject: str, text: str) -> bool:
    host = os.getenv("SMTP_HOST")
    port = safe_int(os.getenv("SMTP_PORT", "465"), 465)
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASS")
    to_addr = os.getenv("SMTP_TO")
    from_addr = os.getenv("SMTP_FROM") or user
    if not host or not user or not password or not to_addr or not from_addr:
        return False
    msg = MIMEText(text, "plain", "utf-8")
    msg["Subject"] = Header(subject, "utf-8")
    msg["From"] = from_addr
    msg["To"] = to_addr
    try:
        if truthy(os.getenv("SMTP_USE_SSL", os.getenv("SMTP_SSL", "true"))):
            server = smtplib.SMTP_SSL(host, port, timeout=20)
        else:
            server = smtplib.SMTP(host, port, timeout=20)
            server.starttls(context=ssl.create_default_context())
        server.login(user, password)
        server.sendmail(from_addr, [to_addr], msg.as_string())
        server.quit()
        return True
    except Exception:
        return False


def parse_channels() -> list[str]:
    raw = (os.getenv("NOTIFY_CHANNELS") or "").strip()
    if raw:
        return [item.strip().lower() for item in raw.split(",") if item.strip()]
    channels = []
    if os.getenv("TELEGRAM_BOT_TOKEN") and os.getenv("TELEGRAM_CHAT_ID"):
        channels.append("telegram")
    if os.getenv("SMTP_HOST") and os.getenv("SMTP_TO"):
        channels.append("email")
    return channels


def is_enabled(env_name: str, default: str = "true") -> bool:
    return truthy(os.getenv(env_name, default))


def main():
    results_dir = sys.argv[1] if len(sys.argv) > 1 else "results"
    account_lookup, expected_total = load_account_lookup()
    manifest = load_manifest(results_dir)
    output_xlsx = resolve_output_xlsx_path(results_dir, manifest)
    raw_records = load_results(results_dir, account_lookup)
    summary_category = str(os.getenv("SUMMARY_CATEGORY") or "").strip()
    if summary_category:
        allowed_prefixes = {
            "老号全干组": {"old"},
            "新号全干组": {"new"},
            "同行不签到组": {"ll", "zh"},
        }.get(summary_category, set())
        raw_records = [
            record for record in raw_records
            if str(record.get("account_category") or "").strip() == summary_category
        ]
        account_lookup = {
            key: username for key, username in account_lookup.items()
            if isinstance(key[0], str) and any(key[0].startswith(prefix) for prefix in allowed_prefixes)
        }
    if os.getenv("EXPECTED_TOTAL") not in (None, ""):
        expected_total = max(0, safe_int(os.getenv("EXPECTED_TOTAL"), 0))
    records = merge_records_with_expected(raw_records, account_lookup, target_date_text(manifest))
    message, summary = build_message(records, manifest, expected_total)

    channels = parse_channels()
    send_tg_text = is_enabled("TELEGRAM_SEND_TEXT", "true")
    send_tg_xlsx = is_enabled("TELEGRAM_SEND_XLSX", "true")
    generate_xlsx = send_tg_xlsx or is_enabled("GENERATE_XLSX", "false")

    if generate_xlsx:
        write_xlsx(output_xlsx, records)

    sent = False
    if "telegram" in channels:
        if send_tg_text:
            sent = send_telegram_message(message) or sent
        if send_tg_xlsx:
            if not os.path.exists(output_xlsx):
                write_xlsx(output_xlsx, records)
            sent = send_telegram_document(output_xlsx) or sent
    if "email" in channels or "smtp" in channels:
        subject = f"{target_date_text(manifest)} 签到汇总"
        sent = send_email(subject, message) or sent

    print(message)
    print(
        f"[summary] total={summary['total']} success={summary['success']} "
        f"sent={'yes' if sent else 'no'} tg_text={'on' if send_tg_text else 'off'} "
        f"tg_xlsx={'on' if send_tg_xlsx else 'off'} xlsx_generated={'yes' if os.path.exists(output_xlsx) else 'no'}"
    )

    if truthy(os.getenv("FAIL_ON_FAILURE", "false")) and summary["problem_count"] > 0:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
